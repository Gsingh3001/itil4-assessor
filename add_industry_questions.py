"""
add_industry_questions.py
─────────────────────────────────────────────────────────────────────────────
Adds an `industry_overlay` column to every existing row in question-bank.xlsx
(default "ALL") and appends industry-specific questions for all 14 industries.

Run once from the repo root:
    python add_industry_questions.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
from copy import copy

XLSX_PATH = "public/question-bank.xlsx"

# ── Colour fills per industry ────────────────────────────────────────────────
FILLS = {
    "Banking & Financial Services":  PatternFill("solid", fgColor="C6EFCE"),  # green
    "Insurance":                      PatternFill("solid", fgColor="FFEB9C"),  # amber
    "Healthcare & Life Sciences":     PatternFill("solid", fgColor="FFC7CE"),  # red
    "Retail & Consumer":              PatternFill("solid", fgColor="BDD7EE"),  # blue
    "Manufacturing":                  PatternFill("solid", fgColor="E2EFDA"),  # olive
    "Telecommunications":             PatternFill("solid", fgColor="FCE4D6"),  # peach
    "Energy & Utilities":             PatternFill("solid", fgColor="D9E1F2"),  # lavender
    "Public Sector / Government":     PatternFill("solid", fgColor="EDEDED"),  # grey
    "Media & Entertainment":          PatternFill("solid", fgColor="FFF2CC"),  # yellow
    "Technology":                     PatternFill("solid", fgColor="DDEBF7"),  # sky
    "Professional Services":          PatternFill("solid", fgColor="E9D7FE"),  # purple
    "Airlines":                       PatternFill("solid", fgColor="DDEEFF"),  # light blue
    "Logistics":                      PatternFill("solid", fgColor="D6FCE5"),  # mint
    "Other":                          PatternFill("solid", fgColor="F5F5F5"),  # white-grey
}

# ── Helper ───────────────────────────────────────────────────────────────────
_id_counter = {}

def q(practice_id, practice_name, group, level, dim, text, hint="", evidence="", guidance="",
      fup_p="", fup_n="", mandatory="N", industry="ALL"):
    key = f"{practice_id}_{industry}_{level}"
    _id_counter[key] = _id_counter.get(key, 0) + 1
    n = _id_counter[key]
    ind_tag = industry[:4].lower().replace(" ","_").replace("&","") if industry != "ALL" else "all"
    qid = f"{practice_id[:6]}_{ind_tag}_{level[:3]}_{n:02d}"
    return [
        practice_id, practice_name, group, level, 999 + n, qid,
        text, hint, evidence, dim, fup_p, fup_n, mandatory, guidance, industry
    ]

# ── Industry-specific question definitions ──────────────────────────────────
#
# Format: q(practice_id, practice_name, group, level, dim, text, hint, evidence,
#            guidance, fup_p, fup_n, mandatory, industry)
# Dimensions: PE=Process Existence, PC=Process Consistency,
#             MM=Measurement Maturity, CI=Continuous Improvement, TI=Tool Integration
# Levels: beginner | practitioner | expert
# ─────────────────────────────────────────────────────────────────────────────

INDUSTRY_QUESTIONS = []

# ═══════════════════════════════════════════════════════════════════════════════
# 1. BANKING & FINANCIAL SERVICES
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Banking & Financial Services"

# Incident Management
INDUSTRY_QUESTIONS += [
    q("incident_mgmt","Incident Management","Service Management","beginner","PE",
      "Does your IT incident management process map to the bank's operational resilience obligations under PRA SS1/21 or equivalent central-bank guidance?",
      "PS rules require firms to identify important business services and set impact tolerances.",
      "Policy document referencing PS/SS1 resilience requirements; IBS mapping artefact.",
      "Absence of alignment indicates a regulatory gap; partial = referenced but not fully mapped.",
      "Which business services have been formally designated as important?",
      "Is there awareness of operational resilience regulatory requirements?",
      "Y", IND),
    q("incident_mgmt","Incident Management","Service Management","practitioner","PC",
      "Are P1/P2 incidents categorised using the bank's own impact taxonomy that considers real-time payment system availability, SWIFT messaging, and core-banking uptime?",
      "Financial-services incidents have direct monetary and regulatory consequences if payment rails go down.",
      "Incident priority matrix with explicit mention of payment/core-banking services.",
      "Best = taxonomy in tool; partial = documented but not enforced; none = generic IT priorities.",
      "How is the impact on real-time payments measured during a major incident?",
      "Does the team understand why payments availability is classified differently?",
      "N", IND),
    q("incident_mgmt","Incident Management","Service Management","expert","CI",
      "Does the major incident review process include root-cause analysis linked back to DORA (Digital Operational Resilience Act) reporting obligations, and are lessons learned shared with the board risk committee?",
      "DORA mandates classification and reporting of ICT-related incidents to regulators.",
      "Board risk committee minutes or DORA incident register; evidence of lessons-learned loop.",
      "Expert level: systematic DORA-linked RCA with board visibility.",
      "How are DORA incidents classified versus internal incidents?",
      "Are major incident findings reported only to IT management?",
      "N", IND),
]

# Change Enablement
INDUSTRY_QUESTIONS += [
    q("change_enablement","Change Enablement","Service Management","beginner","PE",
      "Is there a change freeze calendar aligned to financial year-end, month-end processing windows, and regulatory reporting dates?",
      "Banks cannot afford failed changes during period-close or regulatory submission windows.",
      "Change calendar showing blackout periods tied to finance calendar.",
      "Absence = high risk of year-end disruptions.",
      "Which specific dates are protected and how are emergency changes handled during freezes?",
      "Does the team know when the financial year-end blackout periods are?",
      "Y", IND),
    q("change_enablement","Change Enablement","Service Management","practitioner","PC",
      "Are changes to core-banking, payment-switching, or SWIFT infrastructure subject to a separate enhanced-risk assessment that includes regulatory impact and cyber-security review?",
      "Payment-infrastructure changes often trigger PCI-DSS Section 6 controls and regulator notification.",
      "Change record showing regulatory-impact field; evidence of PCI-DSS CAB sign-off for in-scope systems.",
      "Best = mandatory field + automated routing; partial = manual process; none = standard CAB only.",
      "How is PCI-DSS scope assessed for each change to payment systems?",
      "Are payment-system changes reviewed by the same CAB as general IT changes?",
      "N", IND),
    q("change_enablement","Change Enablement","Service Management","expert","TI",
      "Is change success automatically correlated with post-implementation monitoring of transaction throughput, fraud-detection accuracy, and settlement latency via integrated observability tooling?",
      "Automated correlation reduces MTTR and provides evidence for internal audit.",
      "Observability dashboard showing KPIs before/after each change; linked to change record.",
      "Expert: automated PIR metrics; practitioner: manual PIR; beginner: no PIR linkage.",
      "Which observability platform is used and how long is post-change monitoring sustained?",
      "Are change outcomes assessed purely through incident count?",
      "N", IND),
]

# Information Security Management
INDUSTRY_QUESTIONS += [
    q("info_security_mgmt","Information Security Management","General Management","practitioner","PC",
      "Are IT security controls for customer data and payment card data aligned to both ISO 27001 and PCI-DSS, with documented evidence of annual penetration testing and quarterly vulnerability scanning?",
      "Dual compliance is standard practice; banks are typically PCI-DSS Level 1 merchants/service providers.",
      "PCI-DSS RoC or SAQ; ISO 27001 certificate; pentest/vuln scan reports.",
      "Best = both certifications current + evidence; partial = one certification or outdated.",
      "What is the last PCI-DSS assessment date and when is the next scheduled?",
      "Has the security team heard of PCI-DSS requirements?",
      "Y", IND),
    q("info_security_mgmt","Information Security Management","General Management","expert","CI",
      "Does the organisation operate a 24×7 Security Operations Centre (SOC) with SIEM integration that can detect and contain an Advanced Persistent Threat (APT) within the financial sector's expected 1-hour containment SLA?",
      "Regulators (FCA, PRA, ECB TIBER-EU) expect rapid threat containment.",
      "SOC SLA documentation; SIEM alert-to-containment metrics; TIBER-EU or CBEST exercise results.",
      "Expert = measured sub-60-min containment; practitioner = SOC exists but unvalidated; beginner = no SOC.",
      "What was the mean time to contain (MTTC) in the last TIBER/CBEST exercise?",
      "Does the organisation have a SOC at all?",
      "N", IND),
]

# Service Continuity Management
INDUSTRY_QUESTIONS += [
    q("service_continuity_mgmt","Service Continuity Management","Service Management","practitioner","PC",
      "Does the IT service continuity plan for core banking explicitly meet the RTO/RPO requirements set by the regulator (e.g., same-day recovery for systemically important institutions)?",
      "Many regulators mandate <4h RTO for critical payment infrastructures.",
      "ISCP with regulator-defined RTO/RPO; last DR test report showing achieved RTO.",
      "Best = tested and within regulator threshold; partial = documented but not recently tested.",
      "When was the last DR test performed and what was the achieved RTO?",
      "Does a service continuity plan exist at all?",
      "Y", IND),
]

# Risk Management
INDUSTRY_QUESTIONS += [
    q("risk_mgmt","Risk Management","General Management","practitioner","MM",
      "Are IT and operational risks entered into the enterprise risk management (ERM) framework and reported to the Chief Risk Officer (CRO) using the bank's risk appetite statement?",
      "Basel III/IV mandates operational risk capital allocation; IT risks feed this model.",
      "IT risk register entries visible in ERM; OR capital calculation referencing IT risks.",
      "Expert = fully integrated with capital calculations; practitioner = visible in ERM; beginner = siloed IT risk log.",
      "How do IT risk scores influence operational risk capital provisioning?",
      "Is an IT risk register maintained separately from business risk?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 2. INSURANCE
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Insurance"

INDUSTRY_QUESTIONS += [
    q("incident_mgmt","Incident Management","Service Management","beginner","PE",
      "Does the incident management process include a category specifically for claims-processing system outages, given their direct impact on policyholder obligations?",
      "Claims-system downtime can trigger regulatory penalties for delayed settlement.",
      "Incident category list with claims-processing as a named category.",
      "Absence = potential missed SLA obligations to policyholders.",
      "What is the defined SLA for restoring claims-processing systems after an outage?",
      "Does the team know which systems support claims processing?",
      "Y", IND),
    q("incident_mgmt","Incident Management","Service Management","practitioner","MM",
      "Are incident metrics for policy-administration-system (PAS) and claims-management-system (CMS) outages reported to actuarial and compliance teams to inform Solvency II operational risk modelling?",
      "Solvency II Pillar 2 requires operational risk quantification including IT incidents.",
      "Incident dashboard with PAS/CMS metrics; evidence of data flowing to actuarial models.",
      "Expert = automated feed to OR models; practitioner = periodic manual reporting; beginner = no linkage.",
      "How frequently are IT incident metrics shared with the actuarial team?",
      "Are IT incident statistics shared outside of the IT department?",
      "N", IND),
    q("change_enablement","Change Enablement","Service Management","practitioner","PC",
      "Are changes to underwriting-rule engines and pricing algorithms subject to model governance review in addition to standard IT change process?",
      "Pricing algorithm changes can have material P&L and regulatory impact; model risk governance is best practice.",
      "Change record showing model-governance sign-off field for pricing/underwriting systems.",
      "Best = mandatory dual approval; partial = process exists but inconsistently applied.",
      "How is actuarial sign-off obtained for pricing model changes?",
      "Are pricing changes treated as standard IT changes?",
      "N", IND),
    q("risk_mgmt","Risk Management","General Management","practitioner","PC",
      "Are third-party IT supplier risks assessed against insurance-specific considerations such as data access to policyholder PII and Lloyd's of London minimum standards?",
      "Insurance is heavily outsourced; regulators require formal third-party risk management.",
      "Supplier risk register with insurance-specific risk criteria; Lloyd's Minimum Standards evidence if applicable.",
      "Expert = tiered supplier assurance aligned to LMX/Lloyd's; practitioner = general supplier risk; beginner = no formal process.",
      "Which suppliers have access to policyholder data and how are they risk-rated?",
      "Is there a supplier risk register?",
      "N", IND),
    q("service_continuity_mgmt","Service Continuity Management","Service Management","beginner","PE",
      "Does the IT service continuity plan specifically address catastrophic event scenarios (CAT events) where claims volume could spike 10× within 24 hours?",
      "Major weather events or natural disasters create sudden, extreme demand on claims systems.",
      "ISCP scenario documentation including CAT event surge capacity plan.",
      "Absence = material risk during large-loss events.",
      "What is the maximum surge capacity for claims-processing systems and how was this validated?",
      "Does a continuity plan exist for claims systems?",
      "Y", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 3. HEALTHCARE & LIFE SCIENCES
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Healthcare & Life Sciences"

INDUSTRY_QUESTIONS += [
    q("incident_mgmt","Incident Management","Service Management","beginner","PE",
      "Is there a defined critical-incident category for Electronic Health Record (EHR) and clinical-systems outages that invokes a clinical downtime procedure to protect patient safety?",
      "EHR downtime can directly endanger patients; clinical downtime procedures (paper fallback) are mandatory.",
      "Incident priority matrix with EHR/clinical systems flagged as safety-critical; clinical downtime SOP.",
      "Absence = patient safety risk and potential CQC / Joint Commission violation.",
      "What is the clinical downtime procedure and who is notified when EHR systems go offline?",
      "Does the team know what a clinical downtime procedure is?",
      "Y", IND),
    q("info_security_mgmt","Information Security Management","General Management","practitioner","PC",
      "Are patient data and electronic Protected Health Information (ePHI) protected under a HIPAA / GDPR dual-compliance framework with documented technical and administrative safeguards?",
      "Healthcare organisations processing EU patient data must satisfy both HIPAA (if US-linked) and GDPR.",
      "HIPAA Security Rule risk analysis; GDPR DPIA for patient data; technical safeguard evidence (encryption, access logs).",
      "Expert = dual-certified + continuous monitoring; practitioner = documented controls; beginner = no formal framework.",
      "When was the last HIPAA risk analysis performed and what findings were remediated?",
      "Are technical safeguards for ePHI documented?",
      "Y", IND),
    q("change_enablement","Change Enablement","Service Management","practitioner","PC",
      "Do changes to clinical applications, medical-device integration (e.g., HL7 FHIR interfaces), and diagnostic systems require clinical safety sign-off using a DCB0160 Clinical Risk Management process or equivalent?",
      "UK DCB0160 / NHS DTAC mandates clinical safety officer sign-off for health IT changes.",
      "Change record with clinical safety officer (CSO) approval field; SCRS/hazard log for affected systems.",
      "Expert = DCB0160 embedded in ITSM tooling; practitioner = manual CSO sign-off; beginner = no clinical review.",
      "Who is the designated Clinical Safety Officer and how are they engaged in the change process?",
      "Are clinical application changes reviewed differently from other IT changes?",
      "Y", IND),
    q("service_continuity_mgmt","Service Continuity Management","Service Management","practitioner","MM",
      "Has the RTO for clinical systems been validated against safe operating timelines defined by clinical leadership, and is this tested at least annually with clinical staff participation?",
      "Clinically acceptable downtime may be measured in minutes, not hours.",
      "RTO sign-off document from Chief Medical Officer or clinical leads; DR test report showing clinical participation.",
      "Best = RTO validated and tested with clinicians; partial = IT-defined RTO only.",
      "When did clinical staff last participate in a continuity exercise and what was the outcome?",
      "Is RTO defined for critical clinical systems?",
      "Y", IND),
    q("supplier_mgmt","Supplier Management","General Management","practitioner","PC",
      "Are IT suppliers who handle patient data assessed against DSPT (NHS Data Security & Protection Toolkit) or equivalent health-sector supplier assurance standards?",
      "NHS DSPT is mandatory for suppliers with access to NHS patient data.",
      "Supplier list with DSPT assertion status; evidence of DSPT review in procurement process.",
      "Expert = all suppliers compliant + monitored; practitioner = assessed but gaps exist; beginner = not assessed.",
      "Which suppliers have outstanding DSPT assertions and what is the remediation plan?",
      "Is supplier DSPT compliance tracked?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 4. RETAIL & CONSUMER
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Retail & Consumer"

INDUSTRY_QUESTIONS += [
    q("capacity_performance_mgmt","Capacity & Performance Management","Service Management","practitioner","MM",
      "Is there a documented peak-trading capacity plan that has been load-tested against Black Friday / Cyber Monday traffic volumes at least 8 weeks in advance?",
      "E-commerce platforms can see 10–20× normal traffic during peak events; failure = direct revenue loss.",
      "Load-test report showing peak simulation results; capacity uplift plan; timeline showing 8-week lead.",
      "Expert = automated elastic scaling with load-tested thresholds; practitioner = manual capacity plan; beginner = reactive.",
      "What was the highest concurrent user load tested and how did it compare to the actual peak?",
      "Is there a capacity plan for peak trading periods?",
      "Y", IND),
    q("incident_mgmt","Incident Management","Service Management","practitioner","PC",
      "During trading-critical periods (e.g., peak sale events), is there a war-room or bridge protocol activating within 5 minutes of a P1 incident affecting the e-commerce platform?",
      "Revenue loss during outages is quantifiable; rapid war-room activation minimises impact.",
      "Major incident procedure with peak-trading protocol section; on-call rota for peak periods.",
      "Expert = automated bridge activation; practitioner = defined protocol; beginner = ad-hoc response.",
      "How long did it take to activate the war-room in the most recent peak-period incident?",
      "Is there any special incident procedure for peak trading periods?",
      "Y", IND),
    q("change_enablement","Change Enablement","Service Management","beginner","PE",
      "Is there a trading freeze preventing non-emergency changes to POS systems, e-commerce platforms, and payment gateways during the 4-week peak trading period (typically November–December)?",
      "Failed changes during peak can cause multi-million pound revenue losses.",
      "Change calendar showing retail freeze window; CAB policy enforcing the freeze.",
      "Absence = direct revenue risk; partial = informal freeze without tooling enforcement.",
      "What percentage of P1 incidents last year were caused by changes made during the peak period?",
      "Does the team know when the trading freeze applies?",
      "Y", IND),
    q("service_continuity_mgmt","Service Continuity Management","Service Management","practitioner","PC",
      "Does the service continuity plan include a site-failover scenario for the primary fulfilment and warehouse management systems, with a tested RTO that does not exceed the customer order-promise window?",
      "Customers expect same-day or next-day fulfilment; WMS outages must be recovered within commitment windows.",
      "BCP documentation with WMS failover; DR test report showing achieved RTO vs. order-promise SLA.",
      "Expert = automated WMS failover tested; practitioner = documented plan + tested; beginner = untested.",
      "What is the order-promise SLA and how does it compare to the WMS RTO?",
      "Is there a continuity plan for warehouse management systems?",
      "N", IND),
    q("service_level_mgmt","Service Level Management","Service Management","practitioner","MM",
      "Are SLAs for omnichannel platforms (web, mobile app, in-store kiosk) defined separately, reflecting different customer experience expectations, and reported in real time on a NOC dashboard?",
      "Omnichannel customers expect consistent CX; channel-specific SLAs enable targeted investment.",
      "SLA register with channel-specific targets; NOC/command-centre dashboard showing real-time availability.",
      "Expert = real-time per-channel SLA with auto-alerts; practitioner = defined targets + periodic reporting.",
      "What is the availability SLA for the mobile app versus the web platform?",
      "Are SLAs defined for customer-facing channels?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 5. MANUFACTURING
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Manufacturing"

INDUSTRY_QUESTIONS += [
    q("change_enablement","Change Enablement","Service Management","practitioner","PC",
      "Is there a separate change management process for Operational Technology (OT) and Industrial Control Systems (ICS/SCADA), with mandatory OT security review and plant-floor operational sign-off?",
      "IT change processes are unsuitable for OT; applying them without modification can cause production downtime.",
      "OT change procedure document; evidence of OT security team and plant manager sign-off on recent changes.",
      "Best = separate OT CAB with security review; partial = same process with OT fields; none = no OT process.",
      "Who approves changes to SCADA systems and how is the IT/OT boundary managed?",
      "Is there any change process for SCADA or PLC systems?",
      "Y", IND),
    q("incident_mgmt","Incident Management","Service Management","beginner","PE",
      "Are production-line IT/OT incidents (MES, SCADA, robotics controllers) handled by a process that includes plant-floor operations in the incident bridge and has a defined OEE (Overall Equipment Effectiveness) impact metric?",
      "Manufacturing incidents directly affect OEE; plant operations must be part of the response.",
      "Incident procedure showing OT/MES category with operations notification; OEE impact field in incident record.",
      "Absence = delayed production recovery; partial = IT process used without OT-specific considerations.",
      "How quickly is the plant operations manager notified of a SCADA or MES outage?",
      "Does the incident team know what OEE means and how to contact plant operations?",
      "Y", IND),
    q("service_continuity_mgmt","Service Continuity Management","Service Management","practitioner","PC",
      "Does the service continuity plan address OT network segmentation failures and ransomware scenarios that could cause production line shutdown, with a tested manual-fallback procedure?",
      "Ransomware attacks on manufacturing OT (e.g., NotPetya) caused billions in losses industry-wide.",
      "OT-specific BCP scenario; evidence of OT ransomware tabletop exercise; manual fallback documented.",
      "Expert = annual OT-specific DR test; practitioner = documented OT BCP; beginner = IT-only BCP.",
      "When was a ransomware attack on the OT network last simulated in a tabletop exercise?",
      "Does the continuity plan mention OT or production-line systems?",
      "N", IND),
    q("it_asset_mgmt","IT Asset Management","Service Management","practitioner","PC",
      "Are OT assets (PLCs, HMIs, SCADA servers, industrial IoT devices) included in the asset register, with firmware versions tracked and end-of-life status managed in line with ICS-CERT advisories?",
      "OT asset visibility is foundational for both security (CVE patching) and lifecycle management.",
      "Unified asset register showing OT and IT assets; firmware version field; ICS-CERT advisory cross-reference.",
      "Expert = automated OT discovery + advisory correlation; practitioner = manual OT register; beginner = IT assets only.",
      "How many OT assets are currently end-of-life and what is the remediation plan?",
      "Is there any register of OT or plant-floor devices?",
      "N", IND),
    q("monitoring_event_mgmt","Monitoring & Event Management","Service Management","practitioner","TI",
      "Is OT/ICS network traffic monitored using passive OT-specific monitoring tools (e.g., Claroty, Dragos, Nozomi) that are separate from IT SIEM and tuned to industrial protocol anomalies?",
      "Standard IT SIEM tools cannot interpret Modbus, DNP3, or Profinet protocols used in OT.",
      "OT monitoring tool deployment evidence; integration with IT SOC; industrial protocol alerting configured.",
      "Expert = OT-IT SOC integration with industrial protocol tuning; practitioner = OT tool deployed; beginner = none.",
      "Which industrial protocols are monitored and what anomalies have been detected in the last 90 days?",
      "Is there any monitoring of the plant-floor network?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 6. TELECOMMUNICATIONS
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Telecommunications"

INDUSTRY_QUESTIONS += [
    q("availability_mgmt","Availability Management","Service Management","practitioner","MM",
      "Are network availability metrics (voice, broadband, mobile data) reported at the network-element level in real time, with automated thresholds triggering NOC escalation before customer impact?",
      "Telecoms SLAs are often regulated (Ofcom/FCC); proactive monitoring prevents regulatory breach.",
      "NOC dashboard with element-level availability; automated threshold alerting evidence.",
      "Expert = proactive degradation detection; practitioner = reactive monitoring; beginner = customer-reported.",
      "What is the mean customer minutes lost (CML) metric and how does it trend over the last 12 months?",
      "Is network availability monitored proactively?",
      "Y", IND),
    q("incident_mgmt","Incident Management","Service Management","practitioner","PC",
      "Do major network outages follow an incident management process that includes mandatory regulatory reporting to the national communications authority within the required timeframe (e.g., 24 hours under EU NIS2)?",
      "EU NIS2 and national regulations mandate incident notification for essential services including telecoms.",
      "Major incident SOP with regulatory notification step; evidence of NIS2 reports filed.",
      "Expert = automated regulatory notification trigger; practitioner = manual process documented; beginner = not covered.",
      "When was the last regulatory incident report filed and what triggered it?",
      "Is there a process for regulatory incident notification?",
      "Y", IND),
    q("change_enablement","Change Enablement","Service Management","practitioner","PC",
      "Are changes to core network components (5G RAN, IMS, PCRF, BGP routing) subject to enhanced CAB review with mandatory network architect sign-off and a service-impact simulation in a network lab?",
      "Misconfigured core network changes can cause nationwide outages affecting millions of subscribers.",
      "CAB policy distinguishing core-network changes; lab simulation evidence for recent core changes.",
      "Expert = mandatory lab simulation + automated rollback; practitioner = enhanced review + architect sign-off.",
      "How is the rollback procedure verified for changes to BGP routing or IMS components?",
      "Are core network changes treated differently from server changes?",
      "Y", IND),
    q("capacity_performance_mgmt","Capacity & Performance Management","Service Management","expert","CI",
      "Is network capacity forecasting driven by ML-based traffic prediction models that incorporate subscriber growth, new service launches (e.g., 5G slices), and sporting or entertainment events, enabling proactive spectrum and bandwidth provisioning?",
      "Reactive capacity management causes congestion during known high-demand events.",
      "Capacity forecast model documentation; evidence of proactive provisioning before major events.",
      "Expert = ML-driven proactive provisioning; practitioner = trend-based manual forecast; beginner = reactive.",
      "What model is used for capacity forecasting and how accurate were predictions for the last major event?",
      "Is there a capacity forecast beyond 90 days?",
      "N", IND),
    q("service_level_mgmt","Service Level Management","Service Management","practitioner","MM",
      "Are SLAs differentiated by service tier (consumer broadband vs. enterprise leased line vs. emergency services) with appropriate penalty and escalation mechanisms aligned to Ofcom / FCC requirements?",
      "Tiered SLA management is standard in telecoms; emergency services have the strictest obligations.",
      "SLA register showing tiered targets; evidence of emergency-services SLA compliance reporting.",
      "Expert = automated SLA measurement with tier differentiation; practitioner = tiered SLAs defined.",
      "What is the uptime guarantee for emergency-services connections versus consumer broadband?",
      "Are SLAs the same for all customer types?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 7. ENERGY & UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Energy & Utilities"

INDUSTRY_QUESTIONS += [
    q("service_continuity_mgmt","Service Continuity Management","Service Management","practitioner","PC",
      "Does the IT service continuity plan address Energy Management System (EMS) and SCADA outages with RTOs aligned to grid stability requirements (typically <15 minutes for critical control systems)?",
      "Grid control system downtime can cascade to power outages affecting millions; regulators mandate strict RTOs.",
      "EMS/SCADA ISCP with sub-15-minute RTO; DR test report demonstrating achieved RTO.",
      "Expert = tested sub-15-min RTO with regulator sign-off; practitioner = documented plan; beginner = no OT BCP.",
      "What is the tested RTO for the Energy Management System and when was this last validated?",
      "Is there a continuity plan specifically for grid control systems?",
      "Y", IND),
    q("info_security_mgmt","Information Security Management","General Management","practitioner","PC",
      "Are OT/ICS cybersecurity controls aligned to NERC CIP (North America) or NIS2 Directive (Europe), with documented evidence of annual NERC CIP compliance audits or NIS2 incident reporting capability?",
      "Energy sector OT is regulated under NERC CIP/NIS2; non-compliance carries significant penalties.",
      "NERC CIP compliance audit report or NIS2 implementation evidence; ICS security control documentation.",
      "Expert = compliant with evidence + continuous monitoring; practitioner = documented controls; beginner = not addressed.",
      "Which NERC CIP standards apply to your organisation and what was the last audit finding?",
      "Is the team aware of NERC CIP or NIS2 obligations?",
      "Y", IND),
    q("change_enablement","Change Enablement","Service Management","practitioner","PC",
      "Are changes to grid-connected OT systems (SCADA, RTUs, smart meter head-end) subject to a separate change process approved by the Grid Control Room Manager and the OT cybersecurity team?",
      "Unauthorised grid-connected changes can destabilise the grid and breach NERC CIP change management standards.",
      "OT change procedure with Grid Control Room approval step; evidence of dual sign-off on recent changes.",
      "Expert = dual sign-off enforced in tooling; practitioner = manual dual approval; beginner = no OT process.",
      "How is the Grid Control Room notified of planned changes to SCADA or RTU systems?",
      "Are there any specific change controls for grid-connected systems?",
      "Y", IND),
    q("incident_mgmt","Incident Management","Service Management","beginner","PE",
      "Is there a defined incident category for grid-cyber incidents that triggers both IT incident management and the organisation's OT security playbook, with mandatory notification to the national grid security team?",
      "Energy sector cyber incidents may have physical consequences; dual-track response is essential.",
      "Incident category list with grid-cyber category; OT security playbook; national team notification procedure.",
      "Absence = risk of delayed response to grid-cyber incidents.",
      "Which government or grid operator must be notified in the event of a cyber incident affecting OT?",
      "Is there a cyber incident category for grid-connected systems?",
      "Y", IND),
    q("monitoring_event_mgmt","Monitoring & Event Management","Service Management","practitioner","TI",
      "Are smart-meter data anomalies, SCADA event alarms, and corporate IT security alerts correlated in a single operational dashboard, enabling cross-domain situational awareness for the control room?",
      "OT/IT convergence in energy requires unified visibility to detect coordinated cyber-physical attacks.",
      "Unified dashboard screenshot; evidence of OT/IT event correlation rules.",
      "Expert = automated OT/IT correlation with AI-based anomaly detection; practitioner = unified dashboard; beginner = siloed.",
      "What is an example of an OT/IT correlated alert that was detected and responded to?",
      "Are OT alarms visible alongside IT events?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 8. PUBLIC SECTOR / GOVERNMENT
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Public Sector / Government"

INDUSTRY_QUESTIONS += [
    q("info_security_mgmt","Information Security Management","General Management","beginner","PE",
      "Has the organisation achieved or maintained Cyber Essentials Plus certification (or equivalent national standard, e.g., FedRAMP Moderate in the US), as required for handling government data?",
      "Cyber Essentials is a UK government baseline; FedRAMP is US federal minimum; both are often mandatory.",
      "Current Cyber Essentials Plus certificate or FedRAMP authorisation letter.",
      "Absence = inability to hold government contracts or process OFFICIAL data.",
      "When does the current Cyber Essentials Plus certification expire and who manages the renewal?",
      "Has the organisation heard of Cyber Essentials?",
      "Y", IND),
    q("change_enablement","Change Enablement","Service Management","practitioner","PC",
      "Do changes to citizen-facing digital services undergo assessment under the Government Service Standard (GOV.UK) or equivalent, including accessibility (WCAG 2.2) and security review before release?",
      "GDS Service Standard and Accessibility Regulations are legal obligations for public sector digital services.",
      "Change record with GOV.UK Service Standard checklist; accessibility assessment evidence.",
      "Expert = automated WCAG check in CI/CD; practitioner = manual checklist; beginner = no specific review.",
      "How is WCAG 2.2 compliance verified for new or updated citizen-facing features?",
      "Are changes to citizen-facing services reviewed for accessibility?",
      "Y", IND),
    q("service_level_mgmt","Service Level Management","Service Management","practitioner","MM",
      "Are SLAs for citizen-facing services (e.g., benefits portals, tax filing systems) aligned to central government performance standards and reported publicly via a digital service dashboard?",
      "Government Digital Service guidelines require transparency in service performance.",
      "Performance dashboard link or screenshot; service standard compliance evidence.",
      "Expert = public dashboard with real-time metrics; practitioner = internal SLA reporting; beginner = no SLAs defined.",
      "Are service performance metrics currently published publicly and what do they show?",
      "Are service performance metrics reported anywhere?",
      "N", IND),
    q("supplier_mgmt","Supplier Management","General Management","practitioner","PC",
      "Are all ICT suppliers assessed against government procurement frameworks (e.g., G-Cloud, Crown Commercial Service, DigiMarket) with security questionnaires based on NCSC supply chain guidance?",
      "Government supply-chain security is a critical national security concern; NCSC guidance is authoritative.",
      "Supplier list with G-Cloud/CCS framework status; NCSC supply-chain security assessment evidence.",
      "Expert = continuous supplier monitoring + NCSC aligned; practitioner = assessed at procurement; beginner = not assessed.",
      "How many active suppliers are outside a government-approved framework and what is the risk mitigation?",
      "Are government procurement frameworks used for IT supplier selection?",
      "Y", IND),
    q("risk_mgmt","Risk Management","General Management","practitioner","PC",
      "Are IT risks categorised using the HMG Security Classification levels (OFFICIAL, SECRET, TOP SECRET) to ensure appropriate security controls and escalation paths for data-handling incidents?",
      "Government data classification drives IT risk and security control requirements.",
      "Risk register showing classification-aligned risk categories; security controls mapped to classification levels.",
      "Expert = automated classification enforcement; practitioner = documented mapping; beginner = no classification framework.",
      "How does data classification influence the risk rating and treatment approach for IT risks?",
      "Does the IT risk register reference security classification levels?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 9. MEDIA & ENTERTAINMENT
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Media & Entertainment"

INDUSTRY_QUESTIONS += [
    q("capacity_performance_mgmt","Capacity & Performance Management","Service Management","practitioner","MM",
      "Is CDN (Content Delivery Network) capacity for streaming platforms pre-scaled ahead of major content releases (e.g., blockbuster premieres, live sports) using viewership forecasting models?",
      "Streaming platforms face extreme demand spikes at premiere moments; reactive scaling causes buffering.",
      "CDN capacity plan for upcoming major releases; viewership forecast vs. actual for previous events.",
      "Expert = automated CDN pre-scaling based on ML forecasts; practitioner = manual pre-scaling; beginner = reactive.",
      "What was the peak concurrent viewers for the last major live event and how was CDN capacity prepared?",
      "Is there a capacity plan for major content release events?",
      "Y", IND),
    q("incident_mgmt","Incident Management","Service Management","practitioner","PC",
      "During live broadcast events, is there a dedicated on-call incident commander with authority to invoke failover procedures within 30 seconds of a broadcast system failure?",
      "Live broadcast failures are immediately visible to millions; decision authority must be pre-delegated.",
      "Live-event incident runbook; on-call rota with named incident commander; failover decision tree.",
      "Expert = pre-authorised automatic failover; practitioner = named commander with documented runbook; beginner = ad-hoc.",
      "In the last live broadcast incident, how long did it take to invoke the failover procedure?",
      "Is there a specific incident process for live broadcasts?",
      "Y", IND),
    q("change_enablement","Change Enablement","Service Management","beginner","PE",
      "Is there a change blackout period covering live event broadcast windows, scheduled content premieres, and award shows, enforced in the ITSM tool?",
      "Broadcasting infrastructure changes during live events are extremely high risk.",
      "Change calendar with broadcast blackout periods; CAB policy enforcing the freeze.",
      "Absence = direct broadcast-failure risk during live events.",
      "What is the most recent example of a change being blocked or deferred due to a broadcast blackout?",
      "Does the team know when broadcast blackout windows are active?",
      "Y", IND),
    q("info_security_mgmt","Information Security Management","General Management","practitioner","PC",
      "Are Digital Rights Management (DRM) systems and content-protection controls subject to regular security assessments, with procedures for rapid revocation of compromised DRM keys?",
      "Content piracy and DRM bypass cost the industry billions; key compromise requires rapid response.",
      "DRM security assessment report; key revocation procedure; evidence of revocation drill.",
      "Expert = automated key rotation + revocation tested; practitioner = documented procedure; beginner = no formal process.",
      "How quickly can compromised DRM keys be revoked and how was this last tested?",
      "Is there a procedure for DRM key compromise?",
      "N", IND),
    q("service_continuity_mgmt","Service Continuity Management","Service Management","practitioner","PC",
      "Does the continuity plan include a geo-redundant streaming origin and backup playout capability that can sustain 80% of scheduled content in the event of a primary data-centre failure?",
      "Streaming business continuity requires geographic redundancy; partial service is better than none.",
      "Architecture diagram showing geo-redundancy; DR test report showing 80%+ service during failover.",
      "Expert = fully geo-redundant + tested; practitioner = documented plan; beginner = single-site dependency.",
      "What is the percentage of content that can be served during a primary data-centre failure?",
      "Is there a geo-redundant streaming infrastructure?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 10. TECHNOLOGY
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Technology"

INDUSTRY_QUESTIONS += [
    q("release_mgmt","Release Management","Service Management","practitioner","PC",
      "Are production deployments performed using a CI/CD pipeline with mandatory automated test gates (unit, integration, security SAST/DAST) that prevent deployment on test failure?",
      "Technology companies should practise modern DevOps; gated pipelines are the industry baseline.",
      "CI/CD pipeline configuration showing test gates; evidence of a blocked deployment due to test failure.",
      "Expert = zero-touch deployment with full automated gates; practitioner = CI/CD with some gates; beginner = manual.",
      "What percentage of deployments are fully automated and what is the mean deployment frequency?",
      "Are production deployments done manually?",
      "Y", IND),
    q("change_enablement","Change Enablement","Service Management","practitioner","PC",
      "Is feature flagging used to decouple deployment from release, enabling dark launches and progressive rollouts to canary or beta user segments without requiring a change freeze?",
      "Feature flags are a modern engineering best practice enabling safe, incremental releases.",
      "Feature flag management tool evidence; documented process for progressive rollout.",
      "Expert = full feature-flag programme with metrics; practitioner = feature flags used; beginner = deploy = release.",
      "What percentage of new features are released using feature flags and how are rollback decisions made?",
      "Is the team familiar with feature flags or canary deployments?",
      "N", IND),
    q("monitoring_event_mgmt","Monitoring & Event Management","Service Management","practitioner","TI",
      "Is a full-stack observability platform (metrics, logs, distributed traces) in place, with SLO/SLI dashboards enabling engineering teams to detect latency regressions before customers are impacted?",
      "SRE practices require SLO-based observability; blind-spot monitoring leads to customer-reported incidents.",
      "Observability platform (e.g., Datadog, Grafana, Honeycomb) with SLO dashboard; evidence of SLO breach alerting.",
      "Expert = SLO-driven on-call with error-budget policies; practitioner = observability + SLOs defined; beginner = basic monitoring.",
      "What is the current error-budget burn rate for your primary SLO and what action does it trigger?",
      "Are SLOs and error budgets defined for key services?",
      "N", IND),
    q("incident_mgmt","Incident Management","Service Management","practitioner","PC",
      "Does the incident management process include a post-incident review (PIR) for all P1/P2 incidents within 5 business days, producing a blameless post-mortem with systemic action items tracked to closure?",
      "Blameless post-mortems are an SRE best practice; blame culture suppresses learning.",
      "PIR template; recent P1 post-mortem document with action-item tracking.",
      "Expert = blameless PIR + action closure tracking + shared learning; practitioner = PIR process; beginner = no formal PIR.",
      "What is the average time to complete a post-mortem and close action items after a P1 incident?",
      "Is there a post-incident review process?",
      "Y", IND),
    q("deployment_mgmt","Deployment Management","Technical Management","practitioner","TI",
      "Are all production environments managed as infrastructure-as-code (IaC) using tools such as Terraform or Pulumi, with all changes peer-reviewed and applied via CI/CD pipeline, eliminating manual console access?",
      "IaC is the modern standard for technology companies; it enables auditability and repeatability.",
      "IaC repository; CI/CD pipeline for infrastructure; evidence that manual changes are blocked.",
      "Expert = 100% IaC + drift detection; practitioner = IaC used for most infrastructure; beginner = manual provisioning.",
      "What percentage of infrastructure is managed as IaC and is there a drift-detection process?",
      "Is infrastructure provisioned manually or via IaC?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 11. PROFESSIONAL SERVICES
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Professional Services"

INDUSTRY_QUESTIONS += [
    q("service_level_mgmt","Service Level Management","Service Management","practitioner","MM",
      "Are IT service SLAs for client-facing systems (e.g., client portals, collaboration platforms, billing systems) explicitly agreed in client engagement contracts, with breach reporting obligations?",
      "Professional services firms have contractual IT obligations to clients; SLA breaches can trigger penalties.",
      "Client contract showing IT SLA clauses; SLA breach reporting template.",
      "Expert = automated client SLA reporting in contracts; practitioner = defined client SLAs; beginner = informal.",
      "Which client contracts include IT SLA clauses and when was the last SLA breach report issued?",
      "Are IT SLAs included in client contracts?",
      "N", IND),
    q("knowledge_mgmt","Knowledge Management","General Management","practitioner","PC",
      "Is there a structured knowledge-retention programme for IT systems, capturing tribal knowledge from consultants and architects before they leave engagements, using a maintained knowledge base?",
      "Professional services has high staff turnover; knowledge loss is a significant operational risk.",
      "Knowledge base tool; documented knowledge capture process; evidence of exit interviews.",
      "Expert = mandatory knowledge transfer as part of project closure; practitioner = knowledge base maintained; beginner = ad-hoc.",
      "What percentage of IT knowledge articles were created in the last 6 months by departing staff?",
      "Is there a process for capturing knowledge when staff leave?",
      "N", IND),
    q("info_security_mgmt","Information Security Management","General Management","practitioner","PC",
      "Are client data handling obligations (NDA, data processing agreements, ISO 27001 alignment) formally mapped to IT security controls, with evidence provided to clients on request?",
      "Consulting firms often handle sensitive client data under strict confidentiality obligations.",
      "Client DPA/NDA mapped to technical controls; ISO 27001 certificate or SOC 2 Type II report.",
      "Expert = SOC 2 Type II or ISO 27001 with client-facing evidence pack; practitioner = mapped controls; beginner = NDAs only.",
      "Is there a SOC 2 or ISO 27001 report that can be shared with clients during due diligence?",
      "Are client data handling obligations mapped to IT controls?",
      "Y", IND),
    q("availability_mgmt","Availability Management","Service Management","practitioner","MM",
      "Are availability targets for internal billable-time recording and ERP systems treated as business-critical, with their downtime tracked in terms of lost billable hours and reported to practice leads?",
      "In professional services, timesheet and ERP downtime directly impacts revenue recognition.",
      "Availability report showing timesheet/ERP uptime; business-impact calculation for downtime events.",
      "Expert = automated revenue-impact calculation per availability event; practitioner = tracked and reported.",
      "What is the annual availability of the time-recording system and what is the estimated revenue impact of downtime?",
      "Is downtime for billing systems tracked separately?",
      "N", IND),
    q("supplier_mgmt","Supplier Management","General Management","practitioner","PC",
      "Are cloud-platform suppliers (SaaS collaboration tools, CRM, ERP) assessed for data residency compliance with client contractual obligations (e.g., EU data residency, GDPR Article 46 transfer mechanisms)?",
      "Clients often mandate data residency; SaaS suppliers must be validated before use on client engagements.",
      "Supplier data-residency register; evidence of client-required residency compliance check.",
      "Expert = automated data-residency verification; practitioner = documented supplier data locations; beginner = not assessed.",
      "Which suppliers store data outside the EU and how has this been addressed in client contracts?",
      "Is data residency checked for SaaS suppliers?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 12. AIRLINES
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Airlines"

INDUSTRY_QUESTIONS += [
    q("incident_mgmt","Incident Management","Service Management","beginner","PE",
      "Is there a P1 incident category for Passenger Service System (PSS) and Global Distribution System (GDS) outages, with a defined war-room procedure that includes revenue-management and operations-control centre staff?",
      "PSS outages prevent check-in and boarding; GDS outages stop ticket sales through all channels.",
      "Incident category list with PSS/GDS category; war-room procedure document.",
      "Absence = extended PSS downtime with cascading operational disruption.",
      "What is the SLA for restoring PSS check-in functionality and who owns the war-room bridge?",
      "Does the team know what a PSS or GDS is?",
      "Y", IND),
    q("incident_mgmt","Incident Management","Service Management","practitioner","PC",
      "Does the major incident management process for IROPs (Irregular Operations) include automated passenger impact assessment (number of affected pax, connection misses) feeding the Disruption Management System?",
      "IROPs cost airlines millions daily; automated impact assessment enables faster decision-making.",
      "IROPs procedure with passenger-impact automation; Disruption Management System integration evidence.",
      "Expert = automated DMS feed during incidents; practitioner = defined IROPs process; beginner = manual.",
      "How long does it take to determine the full passenger impact during a major PSS incident?",
      "Is there a specific process for IROPs management?",
      "N", IND),
    q("change_enablement","Change Enablement","Service Management","practitioner","PC",
      "Are changes to safety-critical IT systems (Flight Operations System, Weight & Balance, Electronic Flight Bags, ACARS ground infrastructure) subject to EASA Part-IS or equivalent aviation authority IT security sign-off before deployment?",
      "EASA Part-IS mandates Information Security Management for ATM/ANS and operators; safety-critical systems require regulatory alignment.",
      "Change procedure with EASA Part-IS safety assessment field; evidence of safety officer approval.",
      "Expert = mandatory EASA Part-IS impact assessment in CAB tooling; practitioner = manual safety sign-off.",
      "Who is the designated EASA Part-IS compliance officer and how are they engaged in the change process?",
      "Are safety-critical systems differentiated from standard IT in the change process?",
      "Y", IND),
    q("change_enablement","Change Enablement","Service Management","beginner","PE",
      "Is there a change blackout covering the IATA summer and winter schedule change windows (last Sunday of March and October) when global flight schedules are updated in all reservation systems?",
      "Schedule change day causes extremely high load on PSS; failed changes can corrupt global inventory.",
      "Change calendar showing IATA schedule change blackout; CAB policy enforcing the freeze.",
      "Absence = risk of reservation database corruption during global schedule synchronisation.",
      "When was the last IATA schedule change blackout and what emergency-change requests were handled?",
      "Does the team know when the IATA schedule change windows occur?",
      "Y", IND),
    q("service_continuity_mgmt","Service Continuity Management","Service Management","practitioner","PC",
      "Does the IT service continuity plan include a manual check-in fallback procedure (paper-based or offline mode) with gate agent training, and is this tested annually in alignment with safety drills?",
      "Aviation regulations require airlines to be able to operate without IT in an emergency.",
      "Manual check-in SOP; evidence of annual manual-process drill; gate agent training records.",
      "Expert = annual drill + trained gate agents + paper stock; practitioner = documented plan; beginner = no fallback.",
      "When was the manual check-in procedure last drilled and what improvements were identified?",
      "Is there a procedure for operating check-in without the PSS?",
      "Y", IND),
    q("info_security_mgmt","Information Security Management","General Management","practitioner","PC",
      "Are Passenger Name Record (PNR) and Advance Passenger Information (API) data handling processes compliant with GDPR PNR Directive and IATA Resolution 830d, with documented data retention and deletion schedules?",
      "PNR/API data is heavily regulated; GDPR PNR Directive sets strict retention limits across EU member states.",
      "PNR data retention policy; evidence of automated deletion after regulatory retention period.",
      "Expert = automated retention enforcement; practitioner = documented policy; beginner = no formal process.",
      "What is the retention period for PNR data and how is automated deletion verified?",
      "Is PNR data retention governed by a formal policy?",
      "Y", IND),
    q("monitoring_event_mgmt","Monitoring & Event Management","Service Management","practitioner","TI",
      "Are ACARS (Aircraft Communications Addressing and Reporting System) ground-infrastructure health events and flight-data transmission failures monitored in the IT NOC alongside standard IT alerts?",
      "ACARS ground failures affect real-time flight operations monitoring; cross-domain visibility is critical.",
      "NOC dashboard showing ACARS ground system monitoring; integration evidence.",
      "Expert = ACARS + IT fully integrated NOC monitoring; practitioner = ACARS monitored; beginner = not monitored by IT.",
      "What is the alert threshold for ACARS message delivery failures and who is notified?",
      "Is ACARS ground infrastructure monitored by the IT team?",
      "N", IND),
    q("service_level_mgmt","Service Level Management","Service Management","practitioner","MM",
      "Are SLAs for check-in kiosk availability, boarding gate systems, and baggage-reconciliation systems defined in terms of departure-process impact (e.g., on-time departure contribution) and reported to Airport Operations?",
      "Airport SLAs must align with on-time performance (OTP) targets; IT is a key contributor to OTP.",
      "SLA register with operational impact metrics; Airport Operations reporting evidence.",
      "Expert = OTP-linked IT SLAs with automatic notification to Ops Control; practitioner = defined + reported.",
      "What is the kiosk availability SLA and how does it relate to the airline's OTP target?",
      "Are IT SLAs defined for airport operational systems?",
      "N", IND),
    q("risk_mgmt","Risk Management","General Management","practitioner","PC",
      "Are cybersecurity risks to safety-critical aviation systems (including supply chain risks in avionics software updates) assessed using the ICAO Annex 17 cybersecurity framework and reported to the Accountable Manager?",
      "ICAO Annex 17 amendment 17 introduced cybersecurity into aviation security standards.",
      "Risk register entries aligned to ICAO Annex 17; Accountable Manager sign-off evidence.",
      "Expert = ICAO Annex 17 aligned risk register with Accountable Manager reporting; practitioner = documented risks.",
      "How are supply-chain cyber risks for avionics software updates assessed and managed?",
      "Is the team aware of ICAO cybersecurity obligations?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 13. LOGISTICS
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Logistics"

INDUSTRY_QUESTIONS += [
    q("service_continuity_mgmt","Service Continuity Management","Service Management","practitioner","PC",
      "Does the service continuity plan address Warehouse Management System (WMS) and Transport Management System (TMS) outages, with manual fallback procedures for continued shipment processing?",
      "WMS/TMS outages stop fulfilment and can cascade to client SLA breaches.",
      "WMS/TMS ISCP with manual fallback; DR test report.",
      "Expert = tested automated failover + manual fallback; practitioner = documented plan; beginner = no WMS BCP.",
      "When was the WMS failover last tested and what was the achieved RTO versus the customer order-promise window?",
      "Is there a continuity plan for warehouse or transport management systems?",
      "Y", IND),
    q("monitoring_event_mgmt","Monitoring & Event Management","Service Management","practitioner","TI",
      "Are IoT telemetry events from fleet-tracking devices and cold-chain sensors integrated into the IT event management platform, with anomaly alerts triggering automated ticket creation?",
      "Cold-chain excursions or fleet deviations require immediate response; manual monitoring is too slow.",
      "IoT integration evidence in ITSM tool; automated ticket creation for cold-chain/fleet alerts.",
      "Expert = ML-based anomaly detection with automated response; practitioner = IoT integrated + auto-ticket.",
      "How many cold-chain excursion events were automatically detected and remediated last quarter?",
      "Is fleet or cold-chain IoT data monitored by the IT team?",
      "N", IND),
    q("incident_mgmt","Incident Management","Service Management","practitioner","PC",
      "Are IT incidents that affect cross-border customs and trade compliance systems (e.g., CDS, AES, or broker APIs) categorised as critical, with an escalation path to the Customs Compliance Manager?",
      "Customs system outages can halt border clearance, causing demurrage and regulatory penalties.",
      "Incident category for customs/trade systems; Customs Compliance Manager in escalation matrix.",
      "Expert = automated regulatory-impact calculation; practitioner = defined category + escalation; beginner = not categorised.",
      "What is the SLA for restoring customs declaration systems and who is the escalation contact?",
      "Are customs or trade compliance systems identified in the incident process?",
      "Y", IND),
    q("it_asset_mgmt","IT Asset Management","Service Management","practitioner","PC",
      "Are mobile devices and handheld terminals used in warehouse operations (scanners, PDAs) managed through an MDM (Mobile Device Management) solution with firmware lifecycle tracking and remote-wipe capability?",
      "Warehouse handheld devices are often forgotten in asset management; lost/outdated firmware = security and ops risk.",
      "MDM enrolment report showing handheld devices; firmware version tracking; remote-wipe test evidence.",
      "Expert = 100% MDM coverage + automated firmware management; practitioner = MDM deployed; beginner = unmanaged.",
      "What percentage of warehouse handhelds are enrolled in MDM and what is the firmware currency rate?",
      "Are warehouse handhelds included in the IT asset register?",
      "N", IND),
    q("service_level_mgmt","Service Level Management","Service Management","practitioner","MM",
      "Are IT SLAs for last-mile delivery systems and carrier-API integration points measured against carrier on-time performance and parcel-tracking accuracy, with direct reporting to the Head of Logistics Operations?",
      "Carrier API failures cause tracking outages; tracking accuracy directly impacts customer NPS.",
      "SLA register with carrier-API and tracking-system targets; Operations reporting dashboard.",
      "Expert = real-time carrier API SLA monitoring with auto-alert to Ops; practitioner = defined + reported.",
      "What is the SLA for carrier API availability and what was the last breach?",
      "Are carrier API SLAs defined and measured?",
      "N", IND),
    q("change_enablement","Change Enablement","Service Management","practitioner","PC",
      "Do changes to route-optimisation algorithms and carrier-integration APIs require business sign-off from Network Planning and a rollback-tested deployment with a minimum 48-hour parallel-run period?",
      "Route-optimisation algorithm changes can increase fuel costs or miss delivery SLAs if not tested properly.",
      "Change procedure with Network Planning sign-off field; parallel-run evidence for last algorithm change.",
      "Expert = A/B tested deployment with automated metric comparison; practitioner = parallel run + sign-off.",
      "When was the last route-optimisation algorithm change deployed and how was the impact measured?",
      "Are algorithm changes treated differently from infrastructure changes?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# 14. OTHER
# ═══════════════════════════════════════════════════════════════════════════════
IND = "Other"

INDUSTRY_QUESTIONS += [
    q("incident_mgmt","Incident Management","Service Management","practitioner","PC",
      "Are incident categories and priority definitions tailored to the organisation's specific business context, with impact criteria that reflect actual revenue, reputation, and regulatory exposure?",
      "Generic IT priorities often don't reflect business impact in non-standard industries.",
      "Incident priority matrix with business-specific impact criteria.",
      "Expert = business-validated priority matrix reviewed annually; practitioner = documented; beginner = generic template.",
      "When were incident priority definitions last reviewed with business stakeholders?",
      "Are incident priorities aligned to business impact in your specific industry?",
      "N", IND),
    q("risk_mgmt","Risk Management","General Management","practitioner","PC",
      "Are IT risks assessed in the context of the organisation's specific industry regulatory environment, with named applicable standards and compliance deadlines tracked in the risk register?",
      "Every industry has specific regulations; generic IT risk registers miss compliance obligations.",
      "Risk register entries with regulatory standard references; compliance deadline tracking.",
      "Expert = automated compliance calendar; practitioner = documented regulatory mapping; beginner = generic risk log.",
      "Which industry-specific regulations are referenced in the IT risk register?",
      "Does the IT risk register reference industry-specific regulatory standards?",
      "N", IND),
    q("service_continuity_mgmt","Service Continuity Management","Service Management","practitioner","PC",
      "Have the RTOs and RPOs for critical IT systems been validated against actual business-continuity requirements specific to the organisation's sector, rather than using generic IT defaults?",
      "Business continuity requirements vary widely by sector; IT defaults often misalign with business needs.",
      "RTO/RPO sign-off document from business owner; sector-specific justification for targets.",
      "Expert = business-validated RTO/RPO with sector benchmarks; practitioner = documented business sign-off.",
      "Who signed off the RTO/RPO targets and what was the business justification?",
      "Were RTO/RPO targets set by IT without business validation?",
      "N", IND),
]

# ═══════════════════════════════════════════════════════════════════════════════
# WRITE TO XLSX
# ═══════════════════════════════════════════════════════════════════════════════

def update_xlsx():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Step 1 — find existing headers and add industry_overlay if missing
    headers = [cell.value for cell in ws[1]]
    if "industry_overlay" not in headers:
        col_idx = len(headers) + 1
        ws.cell(row=1, column=col_idx).value = "industry_overlay"
        headers.append("industry_overlay")
        print(f"  Added 'industry_overlay' header at column {col_idx}")
    else:
        col_idx = headers.index("industry_overlay") + 1
        print(f"  'industry_overlay' header already exists at column {col_idx}")

    # Step 2 — backfill all existing rows with "ALL"
    updated = 0
    for row in ws.iter_rows(min_row=2):
        cell = row[col_idx - 1]
        if cell.value is None or str(cell.value).strip() == "":
            cell.value = "ALL"
            updated += 1
    print(f"  Backfilled {updated} existing rows with 'ALL'")

    # Step 3 — remove any previously written industry rows (dedup)
    rows_to_keep = [ws[1]]  # always keep header
    removed = 0
    for row in ws.iter_rows(min_row=2):
        val = str(row[col_idx - 1].value or "").strip()
        if val == "ALL" or val == "":
            rows_to_keep.append(row)
        else:
            removed += 1
    if removed:
        print(f"  Removing {removed} previously written industry rows (re-generating)")
        # Rebuild sheet
        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = ws.title
        for r in rows_to_keep:
            ws2.append([cell.value for cell in r])
        wb.close()
        wb2.save(XLSX_PATH)
        wb2.close()
        # Reload
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb.active

    # Step 4 — append industry-specific questions
    # Expected column order after adding industry_overlay:
    # practice_id, practice_name, group, competency_level, question_order,
    # question_id, question_text, hint, evidence_prompt, dimension,
    # followup_if_partial, followup_if_no, is_mandatory, scoring_guidance, industry_overlay

    appended = 0
    for row_data in INDUSTRY_QUESTIONS:
        industry = row_data[-1]
        fill = FILLS.get(industry, PatternFill("solid", fgColor="FFFFFF"))
        new_row_idx = ws.max_row + 1
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=new_row_idx, column=col_num, value=value)
            cell.fill = fill
        appended += 1

    print(f"  Appended {appended} industry-specific questions")

    wb.save(XLSX_PATH)
    wb.close()
    print(f"\nDone. Saved to {XLSX_PATH}")
    print(f"Total industry-specific questions: {appended}")

    # Summary by industry
    from collections import Counter
    by_industry = Counter(r[-1] for r in INDUSTRY_QUESTIONS)
    for ind, cnt in sorted(by_industry.items()):
        print(f"  {ind}: {cnt} questions")


if __name__ == "__main__":
    print(f"Processing {XLSX_PATH} ...")
    update_xlsx()
